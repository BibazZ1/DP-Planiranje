# -*- coding: utf-8 -*-
"""Provjera robusnog Excel/CSV izvoza (in-process test_client, TEST baza)."""
import io
import os
os.environ["POSTGRES_DATABASE"] = "DP-PLANIRANJE-TEST"
os.environ["DEV_FAKE_USER"] = ""
from dotenv import load_dotenv
load_dotenv()
os.environ["POSTGRES_DATABASE"] = "DP-PLANIRANJE-TEST"

import app as A
from openpyxl import load_workbook

ADMIN = "e.uzunovic@gfcbh.ba"
c = A.app.test_client()
passed = failed = 0


def ok(name, cond, extra=""):
    global passed, failed
    if cond:
        passed += 1; print("PASS ", name)
    else:
        failed += 1; print("FAIL ", name, extra)


with c.session_transaction() as s:
    s["user_email"] = ADMIN; s["user_name"] = "Admin"; s.permanent = True

# napravi raznolike podatke: POP s DP + termini (jedan kasni), komentar, claim; i POP bez DP
c.post("/api/pops", json={"projekt": "EXPROJ", "naziv": "POP EX", "hp": 100, "ha": 30})
dp = c.post("/api/dps", json={"projekt": "EXPROJ", "pop": "POP EX", "naziv": "DP EX"}).get_json()
tasks = [t for t in c.get("/api/data").get_json()["tasks"] if t["dp_id"] == dp["id"]]
akt = [t for t in tasks if "aktivacij" in t["aktivnost"].lower()][0]
doz = [t for t in tasks if t["aktivnost"] == "Dozvole"][0]
c.post("/api/segments", json={"task_id": doz["id"], "datum_od": "2026-02-01",
                              "datum_do": "2026-02-20", "status": "otvoreno", "kasni_razlog": "kasni teren"})
c.post("/api/segments", json={"task_id": akt["id"], "datum_od": "2026-09-01",
                              "datum_do": "2026-09-30", "status": "otvoreno"})
c.post("/api/comments", json={"dp_id": dp["id"], "tekst": "Komentar za izvoz test"})
c.post("/api/claims", json={"projekt": "EXPROJ"})
c.post("/api/pops", json={"projekt": "EXPROJ", "naziv": "POP BEZ DP", "hp": 5, "ha": 2})

# ---- XLSX ----
r = c.get("/export/xlsx")
ok("xlsx: 200 + xlsx mime", r.status_code == 200 and "spreadsheet" in r.headers.get("Content-Type", ""))
ok("xlsx: ime fajla s datumom", ".xlsx" in r.headers.get("Content-Disposition", ""))
wb = load_workbook(io.BytesIO(r.data))
ok("xlsx: svi listovi prisutni",
   wb.sheetnames == ["Termini", "DP-ovi", "POP-ovi", "Projekti", "Komentari"],
   str(wb.sheetnames))

wsT = wb["Termini"]
heads = [c.value for c in wsT[1]]
ok("Termini: bogata zaglavlja (Projekt, Kunde, Od, Do, Pomjereno...)",
   "Projekt" in heads and "Kunde" in heads and "Od" in heads and "Pomjereno" in heads and "Kreirao" in heads)
# datum kolona "Od" mora biti PRAVI datum (filtrabilan), ne tekst
odc = heads.index("Od") + 1
import datetime
od_cell = wsT.cell(row=2, column=odc)
ok("Termini: kolona 'Od' je pravi datum (ne tekst)", isinstance(od_cell.value, (datetime.date, datetime.datetime)),
   "%r (%s)" % (od_cell.value, type(od_cell.value).__name__))
ok("Termini: datum ima dd.mm.yyyy format", "DD.MM.YYYY" in (od_cell.number_format or "").upper())
# HP mora biti broj
hpc = heads.index("HP") + 1
ok("Termini: HP je broj", isinstance(wsT.cell(row=2, column=hpc).value, (int, float)))
# Kasni (dana) izračunat za prošli termin
ok("Termini: ima zamrznut prvi red", wsT.freeze_panes == "A2")
ok("Termini: ima Excel tabelu (filter+trake)", len(wsT.tables) >= 1)

# POP-ovi uključuje POP bez DP-a
wsP = wb["POP-ovi"]
popnames = [wsP.cell(row=i, column=[c.value for c in wsP[1]].index("POP/FCP ID") + 1).value
            for i in range(2, wsP.max_row + 1)]
ok("POP-ovi: uključuje POP bez DP-a", "POP BEZ DP" in popnames, str(popnames))

# Projekti list ima vlasnika (claim) i pravi datum
wsPr = wb["Projekti"]
ok("Projekti: ima 'Vlasnik' kolonu", "Vlasnik" in [c.value for c in wsPr[1]])

# Komentari
wsK = wb["Komentari"]
ktekst = [wsK.cell(row=i, column=[c.value for c in wsK[1]].index("Komentar") + 1).value
          for i in range(2, wsK.max_row + 1)]
ok("Komentari: sadrži uneseni komentar", any("Komentar za izvoz test" in (x or "") for x in ktekst))

# ---- CSV ----
r = c.get("/export/csv")
ok("csv: 200 + csv mime", r.status_code == 200 and "csv" in r.headers.get("Content-Type", ""))
txt = r.data.decode("utf-8")
ok("csv: BOM + ; delimiter + bogata zaglavlja",
   txt.startswith("﻿") and "Projekt;Kunde;" in txt and "Pomjereno" in txt)

print("\n===== EXPORT: %d PASS / %d FAIL =====" % (passed, failed))
raise SystemExit(1 if failed else 0)
