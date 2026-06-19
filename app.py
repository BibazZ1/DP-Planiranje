# -*- coding: utf-8 -*-
"""DP Planiranje - web aplikacija (Flask + PostgreSQL + Azure AD prijava).

Hostuje se kao i ULAZNE-FAKTURE: Docker + Cloudflare tunel, ali u VLASTITOJ
Postgres bazi (DP-PLANIRANJE) i pod vlastitom domenom.
"""
import csv
import datetime
import hashlib
import hmac
import io
import os
import subprocess
import threading
import time

from dotenv import load_dotenv
load_dotenv()  # lokalni razvoj; u Dockeru varijable stižu kroz compose env_file

from flask import Flask, g, jsonify, request, render_template, send_file, session

import database
from database import db, now_iso
from auth import (auth_bp, login_required, api_login_required,
                  PERMANENT_ADMIN, IS_DOCKER)

# --- Azure SQL (samo čitanje - nikad ne pišemo u Azure) ---
SQL_SERVER = os.environ.get("SQL_SERVER", "mihtest.database.windows.net")
SQL_DATABASE = os.environ.get("SQL_DATABASE", "Data Analyst")
SQL_UID = os.environ.get("SQL_UID", "korisnik")
SQL_PWD = os.environ.get("SQL_PWD", "")

app = Flask(__name__)
app.json.ensure_ascii = False
app.config["TEMPLATES_AUTO_RELOAD"] = True
# statika se kešira 7 dana u produkciji (linkovi nose ?v=<git hash> -> cache busting).
# U lokalnom dev modu (DEV_FAKE_USER) NE keširamo: ?v=<git hash> se ne mijenja dok se
# ne commit-a, pa bi keš sakrio nesnimljene izmjene app.js/CSS pri reloadu.
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0 if os.environ.get("DEV_FAKE_USER") else 7 * 24 * 3600

# gzip/brotli kompresija (kao ULAZNE-FAKTURE) — app.js/CSS/JSON višestruko manji
try:
    from flask_compress import Compress
    Compress(app)
except ImportError:
    pass  # lokalni dev bez instaliranog paketa i dalje radi

# --- sesija / kolačići (isti princip kao ULAZNE-FAKTURE) ---
_secret = os.environ.get("API_SECRET_KEY", "")
if not _secret:
    if IS_DOCKER:
        raise ValueError("Nedostaje API_SECRET_KEY u .env!")
    import secrets as _s
    _secret = _s.token_hex(32)  # dev fallback
app.config.update(
    SECRET_KEY=_secret,
    SESSION_COOKIE_NAME="session_dp",
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=IS_DOCKER,   # HTTPS samo u produkciji
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=8 * 3600,
)
app.register_blueprint(auth_bp)

# --- verzija/deploy (isto kao ULAZNE-FAKTURE: /api/deploy-status za DEPLOY.bat) ---
APP_START_TIME = time.time()
APP_DIR = os.path.dirname(os.path.abspath(__file__))


def _git_version():
    """Kratki git hash trenutnog koda (fallback: vrijeme builda)."""
    try:
        return subprocess.check_output(
            ["git", "-c", "safe.directory=/app", "rev-parse", "--short", "HEAD"],
            cwd=APP_DIR, stderr=subprocess.DEVNULL).decode().strip()
    except Exception:
        return datetime.datetime.now().strftime("%Y%m%d-%H%M%S")


APP_GIT_VERSION = _git_version()

STANDARD_AKTIVNOSTI = database.STANDARD_AKTIVNOSTI

TASK_FIELDS = ["aktivnost", "odjel"]
SEG_FIELDS = ["datum_od", "datum_do", "status", "komentar", "eskalacija", "esk_razlog",
              "esk_datum", "kasni_razlog"]
DP_FIELDS = ["pop", "naziv", "lokacija", "voditelj", "hp", "ha", "projekt"]


@app.teardown_appcontext
def close_db(exc):
    database.close_db(exc)


def req_user():
    """Ime efektivnog korisnika (ide u historiju/atribuciju). Uz impersonaciju
    g.user_name je ime korisnika kojeg admin 'gleda', pa atribucija odgovara njemu."""
    u = getattr(g, "user_name", None) or session.get("user_name") or session.get("user_email") or ""
    return u.strip()[:60]


def _int(v):
    try:
        return int(float(v or 0))
    except (TypeError, ValueError):
        return 0


def audit(entity, entity_id, action, polje="", staro="", novo="", label=""):
    database.audit(req_user(), entity, entity_id, action, polje, staro, novo, label)


def log_hist(seg_id, polje, vrijednost):
    db().execute(
        'INSERT INTO seg_history(seg_id,ts,"user",polje,vrijednost) VALUES(%s,%s,%s,%s,%s)',
        (seg_id, now_iso(), req_user(), polje, str(vrijednost)))


def _late_reason_missing(status, datum_do, kasni_razlog):
    """Poslovno pravilo (i na serveru, ne samo UI): termin koji završava PRIJE
    danas a nije 'završeno' MORA imati razlog produženja."""
    today = datetime.date.today().isoformat()
    return (status != "završeno" and datum_do and datum_do < today
            and not (kasni_razlog or "").strip())


# ==================================================================================
# CLAIM: vlasništvo projekta — ko preuzme projekat, samo on (+admin) uređuje
# ==================================================================================
def _claim(projekt):
    if not projekt:
        return None
    return db().execute(
        "SELECT owner_email, owner_name FROM project_claims WHERE projektname=%s",
        (projekt,)).fetchone()


def _require_project_edit(projekt):
    """None ako smije uređivati, inače Flask (response, 403). Admin uvijek smije;
    neclaimovan projekat smije svako; claimovan samo vlasnik."""
    if getattr(g, "is_admin", False):
        return None
    c = _claim(projekt)
    if c is None or (c["owner_email"] or "").lower() == (g.user_email or "").lower():
        return None
    who = c["owner_name"] or c["owner_email"]
    return jsonify({"error": f"Projekt je preuzeo {who} — zatraži pristup.",
                    "locked_by": who, "projekt": projekt}), 403


def _dp_projekt(dp_id):
    r = db().execute("SELECT projekt FROM dps WHERE id=%s", (dp_id,)).fetchone()
    return r["projekt"] if r else None


def _task_projekt(task_id):
    r = db().execute("SELECT d.projekt FROM tasks t JOIN dps d ON d.id=t.dp_id "
                     "WHERE t.id=%s", (task_id,)).fetchone()
    return r["projekt"] if r else None


def _seg_projekt(seg_id):
    r = db().execute("SELECT d.projekt FROM segments s JOIN tasks t ON t.id=s.task_id "
                     "JOIN dps d ON d.id=t.dp_id WHERE s.id=%s", (seg_id,)).fetchone()
    return r["projekt"] if r else None


# --- Sync iz Azure SQL (SREDJENI_Daily) u Postgres, agregirano po projektu ---
_sync_state = {"status": "nikad", "time": None, "error": None, "count": 0}
_sync_lock = threading.Lock()

AZURE_PROJECT_QUERY = """
SELECT [Projektname],
       MAX([Kunde])              AS Kunde,
       MAX([Projectcode Intern]) AS Projectcode,
       SUM([HP])                 AS HP,
       SUM([Trasa(m)])           AS Trasa_m,
       SUM([HA(m)])              AS HA_m,
       SUM([HA stck.])           AS HA_stck,
       SUM([Montage])            AS Montaza,
       MIN([Datum])              AS Datum_od,
       MAX([Datum])              AS Datum_do
FROM [SREDJENI_Daily]
WHERE [Projektname] IS NOT NULL AND LTRIM(RTRIM([Projektname])) <> ''
GROUP BY [Projektname]
"""

# Isto, ali po DANU (Datum) — omogućava sumiranje HP/Trasa/HA/Montaže u rasponu Datum od/do
AZURE_DAILY_QUERY = """
SELECT [Projektname],
       CAST([Datum] AS DATE)     AS Datum,
       SUM([HP])                 AS HP,
       SUM([Trasa(m)])           AS Trasa_m,
       SUM([HA(m)])              AS HA_m,
       SUM([HA stck.])           AS HA_stck,
       SUM([Montage])            AS Montaza
FROM [SREDJENI_Daily]
WHERE [Projektname] IS NOT NULL AND LTRIM(RTRIM([Projektname])) <> ''
  AND [Datum] IS NOT NULL
GROUP BY [Projektname], CAST([Datum] AS DATE)
"""


def sync_projects_from_azure():
    """Povuče zbirne podatke po projektu iz Azure SQL (samo SELECT) u Postgres."""
    if not _sync_lock.acquire(blocking=False):
        return  # sync već u toku
    try:
        _sync_state.update(status="u toku", error=None)
        import pyodbc
        cs = (
            "DRIVER={ODBC Driver 18 for SQL Server};"
            f"SERVER={SQL_SERVER};DATABASE={SQL_DATABASE};"
            f"UID={SQL_UID};PWD={SQL_PWD};"
            "Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30;"
            "ApplicationIntent=ReadOnly"
        )
        az = pyodbc.connect(cs, readonly=True)
        rows = az.cursor().execute(AZURE_PROJECT_QUERY).fetchall()
        # dnevni razrez je "nice-to-have" — ako padne (npr. kolona Datum nedostaje),
        # NE smije srušiti glavni sync projekata
        daily, daily_err = [], None
        try:
            daily = az.cursor().execute(AZURE_DAILY_QUERY).fetchall()
        except Exception as de:
            daily_err = str(de)
        az.close()

        now = now_iso()
        con = database.connect()
        cur = con.cursor()
        cur.executemany(
            "INSERT INTO projects(projektname,kunde,projectcode,hp,trasa_m,ha_m,ha_stck,"
            " montaza,datum_od,datum_do,synced_at) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) "
            "ON CONFLICT(projektname) DO UPDATE SET kunde=EXCLUDED.kunde,"
            " projectcode=EXCLUDED.projectcode, hp=EXCLUDED.hp, trasa_m=EXCLUDED.trasa_m,"
            " ha_m=EXCLUDED.ha_m, ha_stck=EXCLUDED.ha_stck, montaza=EXCLUDED.montaza,"
            " datum_od=EXCLUDED.datum_od, datum_do=EXCLUDED.datum_do,"
            " synced_at=EXCLUDED.synced_at",
            [(r[0].strip(), (r[1] or "").strip(), (r[2] or "").strip(),
              float(r[3] or 0), float(r[4] or 0), float(r[5] or 0), float(r[6] or 0),
              float(r[7] or 0), str(r[8] or ""), str(r[9] or ""), now) for r in rows])
        con.commit()   # projekte commituj ODMAH — ne ovise o dnevnom razrezu
        # po-dan razrez (puna zamjena svaki sync) -> Datum od/do filter može sumirati raspon;
        # odvojen commit + try da greška ovdje ne poništi gornji (projektni) sync
        if daily_err is None:
            try:
                cur.execute("DELETE FROM project_daily")
                cur.executemany(
                    "INSERT INTO project_daily(projektname,datum,hp,trasa_m,ha_m,ha_stck,montaza) "
                    "VALUES(%s,%s,%s,%s,%s,%s,%s)",
                    [(d[0].strip(), str(d[1] or "")[:10],
                      float(d[2] or 0), float(d[3] or 0), float(d[4] or 0),
                      float(d[5] or 0), float(d[6] or 0)) for d in daily])
                con.commit()
            except Exception as de:
                con.rollback()
                daily_err = str(de)
        con.close()
        _sync_state.update(status="ok", time=now, count=len(rows), daily_error=daily_err)
    except Exception as e:
        _sync_state.update(status="greška", error=str(e))
    finally:
        _sync_lock.release()


# ==================================================================================
# E-MAIL (SMTP kao ULAZNE-FAKTURE) — šalje SAMO u produkciji (Docker),
# u razvoju/testovima samo loguje da ne zatrpava inbox
# ==================================================================================
SMTP_SERVER = os.environ.get("SMTP_SERVER", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587") or 587)
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASS = os.environ.get("SMTP_PASS", "")


def send_mail(to_list, subject, html):
    if not to_list:
        return False
    if not IS_DOCKER:
        print(f"[MAIL-DEV] '{subject}' -> {to_list}")
        return True
    if not (SMTP_SERVER and SMTP_USER and SMTP_PASS):
        print("[MAIL] SMTP nije konfigurisan — preskačem.")
        return False
    import smtplib
    from email.mime.text import MIMEText
    msg = MIMEText(html, "html", "utf-8")
    msg["Subject"] = subject
    msg["From"] = f"DP Planiranje <{SMTP_USER}>"
    msg["To"] = ", ".join(to_list)
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(SMTP_USER, to_list, msg.as_string())
        return True
    except Exception as e:
        print("[MAIL] greška:", e)
        return False


def _emails(only_admins=False):
    con = database.connect()
    cur = con.cursor()
    cur.execute("SELECT email FROM allowed_users" +
                (" WHERE role='admin'" if only_admins else ""))
    out = [r[0] for r in cur.fetchall()]
    con.close()
    return out


_MAIL_TBL = ("<table style='border-collapse:collapse;font-family:sans-serif;"
             "font-size:13px'>{}</table>")


def _mail_rows(rows):
    if not rows:
        return "<tr><td style='padding:4px 10px;color:#888'>—</td></tr>"
    return "".join(
        "<tr>" + "".join(
            f"<td style='padding:4px 10px;border-bottom:1px solid #e2e8f0'>{c if c is not None else ''}</td>"
            for c in r) + "</tr>"
        for r in rows)


def _digest_html():
    """Sedmični pregled: šta kasni, šta počinje ove sedmice, aktivne eskalacije."""
    con = database.connect()
    cur = con.cursor()
    today = datetime.date.today().isoformat()
    week_end = (datetime.date.today() + datetime.timedelta(days=7)).isoformat()
    cur.execute(
        "SELECT d.pop, d.naziv, t.aktivnost, s.datum_do, "
        " (CURRENT_DATE - s.datum_do::date) "
        "FROM segments s JOIN tasks t ON t.id=s.task_id JOIN dps d ON d.id=t.dp_id "
        "WHERE s.status <> 'završeno' AND s.datum_do < %s ORDER BY s.datum_do", (today,))
    late = cur.fetchall()
    cur.execute(
        "SELECT d.pop, d.naziv, t.aktivnost, s.datum_od, s.datum_do "
        "FROM segments s JOIN tasks t ON t.id=s.task_id JOIN dps d ON d.id=t.dp_id "
        "WHERE s.datum_od >= %s AND s.datum_od < %s ORDER BY s.datum_od",
        (today, week_end))
    starting = cur.fetchall()
    cur.execute(
        "SELECT d.pop, d.naziv, t.aktivnost, s.esk_razlog "
        "FROM segments s JOIN tasks t ON t.id=s.task_id JOIN dps d ON d.id=t.dp_id "
        "WHERE s.eskalacija=1")
    esk = cur.fetchall()
    con.close()
    dom = os.environ.get("APP_DOMAIN", "")
    return (f"<h2 style='font-family:sans-serif'>DP Planiranje — sedmični pregled</h2>"
            f"<h3 style='font-family:sans-serif;color:#dc2626'>⏰ Kasni ({len(late)})</h3>"
            + _MAIL_TBL.format(_mail_rows(
                [(p, d, a, dd, f"+{n} dana") for p, d, a, dd, n in late]))
            + f"<h3 style='font-family:sans-serif'>Počinje ove sedmice ({len(starting)})</h3>"
            + _MAIL_TBL.format(_mail_rows(starting))
            + f"<h3 style='font-family:sans-serif;color:#d97706'>Aktivne eskalacije ({len(esk)})</h3>"
            + _MAIL_TBL.format(_mail_rows(esk))
            + f"<p style='font-family:sans-serif'><a href='{dom}'>Otvori DP Planiranje →</a></p>")


def _notify_eskalacija(seg_id):
    """Instant mail adminima kad neko digne eskalaciju (u pozadini)."""
    def run():
        try:
            con = database.connect()
            cur = con.cursor()
            cur.execute(
                "SELECT d.pop, d.naziv, t.aktivnost, s.datum_od, s.datum_do, s.esk_razlog "
                "FROM segments s JOIN tasks t ON t.id=s.task_id "
                "JOIN dps d ON d.id=t.dp_id WHERE s.id=%s", (seg_id,))
            r = cur.fetchone()
            con.close()
            if not r:
                return
            dom = os.environ.get("APP_DOMAIN", "")
            send_mail(_emails(only_admins=True),
                      f"Eskalacija: {r[1]} — {r[2]}",
                      f"<p style='font-family:sans-serif'><b>{r[0]} · {r[1]}</b> — {r[2]} "
                      f"({r[3]} – {r[4]})<br>Razlog: {r[5] or '—'}</p>"
                      f"<p><a href='{dom}'>Otvori DP Planiranje →</a></p>")
        except Exception as e:
            print("[MAIL] eskalacija:", e)
    threading.Thread(target=run, daemon=True).start()


@app.route("/api/health")
def api_health():
    """Healthcheck za Docker — provjerava i bazu."""
    try:
        db().execute("SELECT 1").fetchone()
        return jsonify({"status": "ok", "db": "ok"})
    except Exception as e:
        return jsonify({"status": "error", "db": str(e)[:200]}), 503


# Datoteke čiji se MD5 poredi lokalno vs. kontejner (DEPLOY.bat verifikacija)
_DEPLOY_HASH_FILES = {
    "app.py": os.path.join(APP_DIR, "app.py"),
    "static/app.js": os.path.join(APP_DIR, "static", "app.js"),
    "static/style.css": os.path.join(APP_DIR, "static", "style.css"),
    "templates/index.html": os.path.join(APP_DIR, "templates", "index.html"),
}


@app.route("/api/deploy-status")
def deploy_status():
    """Stanje deploya za DEPLOY.bat: git commit + MD5 ključnih fajlova.

    Isti princip kao ULAZNE-FAKTURE — zaštićeno X-API-KEY (API_SECRET_KEY),
    jer otkriva git hash i checksume fajlova.
    """
    if not hmac.compare_digest(request.headers.get("X-API-KEY", ""), _secret):
        return jsonify({"error": "Unauthorized"}), 401

    result = {
        "version": APP_GIT_VERSION,
        "uptime_seconds": int(time.time() - APP_START_TIME),
        "timestamp": datetime.datetime.now().isoformat(),
    }

    # živi git HEAD (čita se pri zahtjevu): subprocess -> .git/HEAD -> manifest -> verzija
    git_commit, git_branch = None, "main"
    try:
        git_commit = subprocess.check_output(
            ["git", "-c", "safe.directory=/app", "rev-parse", "HEAD"],
            cwd=APP_DIR, stderr=subprocess.DEVNULL).decode().strip()
        git_branch = subprocess.check_output(
            ["git", "-c", "safe.directory=/app", "rev-parse", "--abbrev-ref", "HEAD"],
            cwd=APP_DIR, stderr=subprocess.DEVNULL).decode().strip()
    except Exception:
        pass
    if not git_commit or len(git_commit) < 40:
        try:
            with open(os.path.join(APP_DIR, ".git", "HEAD")) as f:
                head = f.read().strip()
            if head.startswith("ref: "):
                with open(os.path.join(APP_DIR, ".git", head[5:])) as f:
                    git_commit = f.read().strip()
                git_branch = head[5:].split("/")[-1]
            elif len(head) == 40:
                git_commit = head
        except Exception:
            pass
    if not git_commit or len(git_commit) < 40:
        try:
            import json as _json
            with open(os.path.join(APP_DIR, ".deploy_manifest.json")) as f:
                m = _json.load(f)
            if m.get("commit") and len(m["commit"]) == 40:
                git_commit = m["commit"]
                git_branch = m.get("branch", "main")
        except Exception:
            pass
    if not git_commit:
        git_commit = APP_GIT_VERSION
    result["git"] = {"commit": git_commit, "branch": git_branch,
                     "short": git_commit[:7] if len(git_commit) >= 7 else git_commit}

    # živi MD5 ključnih fajlova (dokazuje šta je STVARNO na disku)
    hashes = {}
    for name, path in _DEPLOY_HASH_FILES.items():
        try:
            with open(path, "rb") as f:
                hashes[name] = hashlib.md5(f.read()).hexdigest()
        except Exception:
            hashes[name] = "error"
    result["file_hashes"] = hashes
    return jsonify(result)


@app.route("/")
@login_required
def index():
    return render_template("index.html", auth={
        "email": g.user_email, "name": g.user_name, "is_admin": g.is_admin,
        "impersonating": getattr(g, "impersonating", False),
        "real_email": getattr(g, "real_email", g.user_email),
        "real_name": getattr(g, "real_name", g.user_name),
        "real_is_admin": getattr(g, "real_is_admin", g.is_admin)},
        v=APP_GIT_VERSION)


@app.route("/api/data")
@api_login_required
def api_data():
    dps = [dict(r) for r in db().execute(
        "SELECT d.id, COALESCE(p.naziv, d.pop) AS pop, d.naziv, d.lokacija, "
        " d.voditelj, d.hp, d.ha, COALESCE(NULLIF(p.projekt,''), d.projekt) AS projekt, "
        " d.pop_id "
        "FROM dps d LEFT JOIN pops p ON p.id = d.pop_id ORDER BY pop, d.naziv")]
    pops = [dict(r) for r in db().execute(
        "SELECT * FROM pops ORDER BY projekt, naziv")]
    tasks = [dict(r) for r in db().execute(
        "SELECT id, dp_id, aktivnost, odjel FROM tasks ORDER BY dp_id, id")]
    segments = [dict(r) for r in db().execute(
        "SELECT * FROM segments ORDER BY task_id, datum_od")]
    # samo svježa historija (hovercard) — puna historija ide kroz /api/history;
    # bez limita bi payload rastao neograničeno s mjesecima korištenja
    history = [dict(r) for r in db().execute(
        'SELECT seg_id, ts, "user", polje, vrijednost FROM seg_history '
        "ORDER BY ts DESC, id DESC LIMIT 500")]
    claims = {r["projektname"]: {"email": r["owner_email"], "name": r["owner_name"]}
              for r in db().execute(
                  "SELECT projektname, owner_email, owner_name FROM project_claims")}
    # zadnji komentar po DP-u (za inline prikaz u grupnom redu — hover = puni tekst)
    last_comments = {r["dp_id"]: {"tekst": r["tekst"], "user": r["user"], "ts": r["ts"]}
                     for r in db().execute(
                         'SELECT DISTINCT ON (dp_id) dp_id, tekst, "user", ts '
                         "FROM dp_comments ORDER BY dp_id, id DESC")}
    return jsonify({"dps": dps, "pops": pops, "tasks": tasks, "segments": segments,
                    "history": history, "claims": claims, "last_comments": last_comments})


@app.route("/api/pops", methods=["POST"])
@api_login_required
def add_pop():
    j = request.get_json(force=True)
    naziv = (j.get("naziv") or "").strip()
    projekt = (j.get("projekt") or "").strip()
    rfa = (j.get("rfa") or "").strip()
    if not naziv:
        return jsonify({"error": "naziv je obavezan"}), 400
    if not rfa:
        return jsonify({"error": "RFA datum je obavezan"}), 400
    blk = _require_project_edit(projekt)
    if blk:
        return blk
    ex = db().execute("SELECT id FROM pops WHERE naziv=%s AND projekt=%s",
                      (naziv, projekt)).fetchone()
    if ex:
        return jsonify({"error": "postoji", "id": ex["id"]}), 409
    hp, ha = _int(j.get("hp")), _int(j.get("ha"))
    cur = db().execute(
        "INSERT INTO pops(projekt,naziv,hp,ha,rfa,created_at,created_by) "
        "VALUES(%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (projekt, naziv, hp, ha, rfa, now_iso(), req_user()))
    pop_id = cur.fetchone()["id"]
    audit("pop", pop_id, "kreirano", novo=f"RFA {rfa} · HP {hp} · HA {ha}", label=naziv)
    db().commit()
    return jsonify({"id": pop_id}), 201


@app.route("/api/pops/<int:pop_id>", methods=["PATCH", "DELETE"])
@api_login_required
def edit_pop(pop_id):
    old = db().execute("SELECT * FROM pops WHERE id=%s", (pop_id,)).fetchone()
    if old is None:
        return jsonify({"error": "ne postoji"}), 404
    blk = _require_project_edit(old["projekt"])
    if blk:
        return blk
    if request.method == "DELETE":
        for d_ in db().execute("SELECT id, naziv FROM dps WHERE pop_id=%s",
                               (pop_id,)).fetchall():
            audit("dp", d_["id"], "obrisano", label=d_["naziv"],
                  staro=f'{old["naziv"]} · {d_["naziv"]}')
            db().execute("DELETE FROM dps WHERE id=%s", (d_["id"],))
        db().execute("DELETE FROM pops WHERE id=%s", (pop_id,))
        audit("pop", pop_id, "obrisano", label=old["naziv"], staro=old["naziv"])
        db().commit()
        return "", 204
    j = request.get_json(force=True)
    sets = {}
    if "naziv" in j and (j["naziv"] or "").strip():
        sets["naziv"] = j["naziv"].strip()
    if "rfa" in j:
        sets["rfa"] = (j["rfa"] or "").strip()
    for k in ("hp", "ha"):
        if k in j:
            sets[k] = _int(j[k])
    if sets:
        q = ", ".join(f"{k}=%s" for k in sets)
        db().execute(f"UPDATE pops SET {q} WHERE id=%s", (*sets.values(), pop_id))
        if "naziv" in sets and sets["naziv"] != old["naziv"]:
            # sinhronizuj tekstualni naziv POP-a na svim njegovim DP-ovima
            db().execute("UPDATE dps SET pop=%s WHERE pop_id=%s",
                         (sets["naziv"], pop_id))
        for k, v in sets.items():
            if str(old[k]) != str(v):
                audit("pop", pop_id, "izmjena", polje=k, staro=old[k], novo=v,
                      label=sets.get("naziv", old["naziv"]))
        db().commit()
    return jsonify({"ok": True})


@app.route("/api/dps", methods=["POST"])
@api_login_required
def add_dp():
    j = request.get_json(force=True)
    d = db()
    pop_id = j.get("pop_id") or None
    projekt = (j.get("projekt") or "").strip()
    pop_name = (j.get("pop") or "").strip()
    prow = None
    if pop_id:
        prow = d.execute("SELECT * FROM pops WHERE id=%s", (pop_id,)).fetchone()
    if prow is None and pop_name:
        prow = d.execute("SELECT * FROM pops WHERE naziv=%s AND projekt=%s",
                         (pop_name, projekt)).fetchone()
        if prow is None:
            # POP upisan kao novi naziv -> automatski se kreira pod projektom;
            # RFA datum je obavezan i za ovaj (auto) put kreiranja POP-a
            rfa = (j.get("rfa") or "").strip()
            if not rfa:
                return jsonify({"error": "RFA datum je obavezan za novi POP"}), 400
            cur = d.execute(
                "INSERT INTO pops(projekt,naziv,rfa,created_at,created_by) "
                "VALUES(%s,%s,%s,%s,%s) RETURNING id",
                (projekt, pop_name, rfa, now_iso(), req_user()))
            new_pop_id = cur.fetchone()["id"]
            audit("pop", new_pop_id, "kreirano", novo=f"RFA {rfa}", label=pop_name)
            prow = d.execute("SELECT * FROM pops WHERE id=%s",
                             (new_pop_id,)).fetchone()
    if prow is not None:
        pop_id, pop_name = prow["id"], prow["naziv"]
        projekt = prow["projekt"] or projekt
    blk = _require_project_edit(projekt)
    if blk:
        return blk
    naziv = j.get("naziv", "")
    hp, ha = _int(j.get("hp")), _int(j.get("ha"))
    cur = d.execute(
        "INSERT INTO dps(pop,naziv,lokacija,voditelj,hp,ha,projekt,pop_id) "
        "VALUES(%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (pop_name, naziv, j.get("lokacija", ""), j.get("voditelj", ""),
         hp, ha, projekt, pop_id))
    dp_id = cur.fetchone()["id"]
    audit("dp", dp_id, "kreirano", novo=f"HP {hp} · HA {ha}", label=naziv)
    for akt, odjel in STANDARD_AKTIVNOSTI:
        d.execute("INSERT INTO tasks(dp_id,aktivnost,odjel) VALUES(%s,%s,%s)",
                  (dp_id, akt, odjel))
    d.commit()
    return jsonify({"id": dp_id}), 201


@app.route("/api/dps/<int:dp_id>", methods=["PATCH", "DELETE"])
@api_login_required
def edit_dp(dp_id):
    old = db().execute("SELECT * FROM dps WHERE id=%s", (dp_id,)).fetchone()
    if old is not None:
        blk = _require_project_edit(old["projekt"])
        if blk:
            return blk
    if request.method == "DELETE":
        db().execute("DELETE FROM dps WHERE id=%s", (dp_id,))
        if old is not None:
            audit("dp", dp_id, "obrisano", label=old["naziv"],
                  staro=f'{old["pop"]} · {old["naziv"]}')
        db().commit()
        return "", 204
    j = request.get_json(force=True)
    sets = {k: j[k] for k in DP_FIELDS if k in j}
    if sets:
        q = ", ".join(f"{k}=%s" for k in sets)
        db().execute(f"UPDATE dps SET {q} WHERE id=%s", (*sets.values(), dp_id))
        if old is not None:
            for k, v in sets.items():
                if str(old[k]) != str(v):
                    audit("dp", dp_id, "izmjena", polje=k, staro=old[k], novo=v,
                          label=sets.get("naziv", old["naziv"]))
        db().commit()
    return jsonify({"ok": True})


@app.route("/api/tasks", methods=["POST"])
@api_login_required
def add_task():
    j = request.get_json(force=True)
    blk = _require_project_edit(_dp_projekt(j["dp_id"]))
    if blk:
        return blk
    akt = j.get("aktivnost", "Nova aktivnost")
    cur = db().execute(
        "INSERT INTO tasks(dp_id,aktivnost,odjel) VALUES(%s,%s,%s) RETURNING id",
        (j["dp_id"], akt, j.get("odjel", "")))
    task_id = cur.fetchone()["id"]
    dpr = db().execute("SELECT naziv FROM dps WHERE id=%s", (j["dp_id"],)).fetchone()
    audit("dp", j["dp_id"], "aktivnost dodana", novo=akt,
          label=dpr["naziv"] if dpr else "")
    db().commit()
    return jsonify({"id": task_id}), 201


@app.route("/api/tasks/<int:task_id>", methods=["PATCH", "DELETE"])
@api_login_required
def edit_task(task_id):
    old = db().execute(
        "SELECT t.*, d.naziv AS dp_naziv, d.projekt AS projekt FROM tasks t "
        "JOIN dps d ON d.id=t.dp_id WHERE t.id=%s", (task_id,)).fetchone()
    if old is not None:
        blk = _require_project_edit(old["projekt"])
        if blk:
            return blk
    if request.method == "DELETE":
        db().execute("DELETE FROM tasks WHERE id=%s", (task_id,))
        if old is not None:
            audit("dp", old["dp_id"], "aktivnost obrisana",
                  staro=old["aktivnost"], label=old["dp_naziv"])
        db().commit()
        return "", 204
    j = request.get_json(force=True)
    sets = {k: j[k] for k in TASK_FIELDS if k in j}
    if sets:
        q = ", ".join(f"{k}=%s" for k in sets)
        db().execute(f"UPDATE tasks SET {q} WHERE id=%s", (*sets.values(), task_id))
        if old is not None:
            for k, v in sets.items():
                if str(old[k]) != str(v):
                    audit("dp", old["dp_id"], "izmjena", polje=k,
                          staro=old[k], novo=v, label=old["dp_naziv"])
        db().commit()
    return jsonify({"ok": True})


@app.route("/api/segments", methods=["POST"])
@api_login_required
def add_segment():
    j = request.get_json(force=True)
    blk = _require_project_edit(_task_projekt(j["task_id"]))
    if blk:
        return blk
    # jedna aktivnost = JEDNA traka: drugi termin na istom redu nije dozvoljen
    ex = db().execute("SELECT id FROM segments WHERE task_id=%s",
                      (j["task_id"],)).fetchone()
    if ex:
        return jsonify({"error": "termin već postoji za ovu aktivnost",
                        "id": ex["id"]}), 409
    if _late_reason_missing(j.get("status", "otvoreno"),
                            j.get("datum_do", ""), j.get("kasni_razlog", "")):
        return jsonify({"error": "razlog produženja je obavezan — termin završava prije danas"}), 400
    cur = db().execute(
        "INSERT INTO segments(task_id,datum_od,datum_do,status,komentar,eskalacija,"
        "esk_razlog,esk_datum,kasni_razlog,created_by,orig_od,orig_do) "
        "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (j["task_id"], j["datum_od"], j["datum_do"], j.get("status", "otvoreno"),
         j.get("komentar", ""), 1 if j.get("eskalacija") else 0, j.get("esk_razlog", ""),
         j.get("esk_datum", ""), j.get("kasni_razlog", ""), req_user(),
         j["datum_od"], j["datum_do"]))
    seg_id = cur.fetchone()["id"]
    log_hist(seg_id, "kreirano",
             f"{j['datum_od']} – {j['datum_do']} · {j.get('status', 'otvoreno')}")
    for k in ("komentar", "esk_razlog", "kasni_razlog"):
        if j.get(k):
            log_hist(seg_id, k, j[k])
    db().commit()
    if j.get("eskalacija"):
        _notify_eskalacija(seg_id)
    return jsonify({"id": seg_id}), 201


@app.route("/api/segments/<int:seg_id>", methods=["PATCH", "DELETE"])
@api_login_required
def edit_segment(seg_id):
    blk = _require_project_edit(_seg_projekt(seg_id))
    if blk:
        return blk
    if request.method == "DELETE":
        # zapamti u trajni dnevnik prije brisanja (seg_history nestaje s terminom)
        info = db().execute(
            "SELECT s.datum_od, s.datum_do, t.aktivnost, t.dp_id, d.naziv AS dp_naziv "
            "FROM segments s JOIN tasks t ON t.id=s.task_id JOIN dps d ON d.id=t.dp_id "
            "WHERE s.id=%s", (seg_id,)).fetchone()
        db().execute("DELETE FROM segments WHERE id=%s", (seg_id,))
        if info is not None:
            audit("dp", info["dp_id"], "termin obrisan", polje=info["aktivnost"],
                  staro=f'{info["datum_od"]} – {info["datum_do"]}',
                  label=info["dp_naziv"])
        db().commit()
        return "", 204
    j = request.get_json(force=True)
    if "eskalacija" in j:
        j["eskalacija"] = 1 if j["eskalacija"] else 0
    sets = {k: j[k] for k in SEG_FIELDS if k in j}
    # plan_qty: ručna planska količina termina; "" / null -> NULL (auto raspodjela)
    if "plan_qty" in j:
        v = j["plan_qty"]
        sets["plan_qty"] = None if v is None or str(v).strip() == "" else max(0.0, float(v))
    if sets:
        old = db().execute("SELECT * FROM segments WHERE id=%s", (seg_id,)).fetchone()
        if old is not None and _late_reason_missing(
                sets.get("status", old["status"]),
                sets.get("datum_do", old["datum_do"]),
                sets.get("kasni_razlog", old["kasni_razlog"])):
            return jsonify({"error": "razlog produženja je obavezan — termin završava prije danas"}), 400
        q = ", ".join(f"{k}=%s" for k in sets)
        db().execute(f"UPDATE segments SET {q} WHERE id=%s", (*sets.values(), seg_id))
        if old is not None:
            # historija: zapiši svaku stvarnu promjenu
            for k, v in sets.items():
                if old[k] != v:
                    if k == "eskalacija":
                        log_hist(seg_id, k, "uključena" if v else "isključena")
                    elif k in ("datum_od", "datum_do"):
                        log_hist(seg_id, k, f"{old[k]} → {v}")
                    elif k == "plan_qty":
                        log_hist(seg_id, k, "auto" if v is None else f"{v:g}")
                    else:
                        log_hist(seg_id, k, v or "(obrisano)")
        db().commit()
        # nova eskalacija (0 -> 1) -> instant mail adminima
        if old is not None and sets.get("eskalacija") == 1 and not old["eskalacija"]:
            _notify_eskalacija(seg_id)
    return jsonify({"ok": True})


@app.route("/api/comments")
@api_login_required
def get_comments():
    """Komentari po DP-u (komandni centar u bočnom panelu)."""
    dp_id = request.args.get("dp_id", type=int)
    rows = [dict(r) for r in db().execute(
        'SELECT id, ts, "user", tekst FROM dp_comments '
        "WHERE dp_id=%s ORDER BY id DESC LIMIT 100", (dp_id,))]
    return jsonify({"comments": rows})


# ==================================================================================
# CLAIM endpoints: preuzmi / otpusti projekat, zatraži pristup
# ==================================================================================
@app.route("/api/claims", methods=["POST"])
@api_login_required
def claim_project():
    j = request.get_json(force=True)
    projekt = (j.get("projekt") or "").strip()
    if not projekt:
        return jsonify({"error": "projekt je obavezan"}), 400
    c = _claim(projekt)
    if c and (c["owner_email"] or "").lower() != g.user_email and not g.is_admin:
        return jsonify({"error": f"Već preuzeo {c['owner_name'] or c['owner_email']}",
                        "locked_by": c["owner_name"] or c["owner_email"]}), 403
    db().execute(
        "INSERT INTO project_claims(projektname,owner_email,owner_name,claimed_at) "
        "VALUES(%s,%s,%s,%s) ON CONFLICT(projektname) DO UPDATE SET "
        "owner_email=EXCLUDED.owner_email, owner_name=EXCLUDED.owner_name, claimed_at=EXCLUDED.claimed_at",
        (projekt, g.user_email, g.user_name, now_iso()))
    audit("projekt", 0, "preuzet", novo=g.user_name, label=projekt)
    db().commit()
    return jsonify({"owner_email": g.user_email, "owner_name": g.user_name}), 201


@app.route("/api/claims", methods=["DELETE"])
@api_login_required
def release_project():
    projekt = (request.args.get("projekt") or "").strip()
    c = _claim(projekt)
    if not c:
        return jsonify({"ok": True})
    if (c["owner_email"] or "").lower() != g.user_email and not g.is_admin:
        return jsonify({"error": "Samo vlasnik ili admin može otpustiti projekat"}), 403
    db().execute("DELETE FROM project_claims WHERE projektname=%s", (projekt,))
    audit("projekt", 0, "otpušten", staro=c["owner_name"] or c["owner_email"], label=projekt)
    db().commit()
    return jsonify({"ok": True})


@app.route("/api/claims/request", methods=["POST"])
@api_login_required
def request_project_access():
    j = request.get_json(force=True)
    projekt = (j.get("projekt") or "").strip()
    c = _claim(projekt)
    if not c:
        return jsonify({"error": "Projekat nije preuzet"}), 400
    dom = os.environ.get("APP_DOMAIN", "")
    send_mail([c["owner_email"]],
              f"Zahtjev za pristup projektu: {projekt}",
              f"<p style='font-family:sans-serif'><b>{g.user_name}</b> ({g.user_email}) "
              f"traži pristup uređivanju projekta <b>{projekt}</b>.<br>"
              f"Ako želiš, otpusti projekat u aplikaciji da preuzme uređivanje.</p>"
              f"<p><a href='{dom}'>Otvori DP Planiranje →</a></p>")
    return jsonify({"ok": True, "owner": c["owner_name"] or c["owner_email"]})


@app.route("/api/comments", methods=["POST"])
@api_login_required
def add_comment():
    j = request.get_json(force=True)
    tekst = (j.get("tekst") or "").strip()[:500]
    dp_id = j.get("dp_id")
    if not tekst or not dp_id:
        return jsonify({"error": "prazan komentar"}), 400
    blk = _require_project_edit(_dp_projekt(dp_id))
    if blk:
        return blk
    cur = db().execute(
        'INSERT INTO dp_comments(dp_id,ts,"user",tekst) '
        "VALUES(%s,%s,%s,%s) RETURNING id",
        (dp_id, now_iso(), req_user(), tekst))
    cid = cur.fetchone()["id"]
    audit("dp", dp_id, "komentar", novo=tekst[:120])
    db().commit()
    return jsonify({"id": cid}), 201


@app.route("/api/projects")
@api_login_required
def api_projects():
    rows = [dict(r) for r in db().execute(
        "SELECT * FROM projects ORDER BY projektname")]
    return jsonify({"projects": rows, "sync": _sync_state})


@app.route("/api/projects/totals")
@api_login_required
def api_projects_totals():
    """Σ HP/Trasa/HA/Montaža po projektu, suženo na raspon Datum od/do (project_daily).
    Bez od/do = ukupno (poklapa se s kolonama u 'projects')."""
    od = (request.args.get("od") or "").strip()
    do = (request.args.get("do") or "").strip()
    cond, params = [], []
    if od:
        cond.append("datum >= %s")
        params.append(od)
    if do:
        cond.append("datum <= %s")
        params.append(do)
    where = (" WHERE " + " AND ".join(cond)) if cond else ""
    rows = [dict(r) for r in db().execute(
        "SELECT projektname, SUM(hp) AS hp, SUM(trasa_m) AS trasa_m, "
        "SUM(ha_m) AS ha_m, SUM(ha_stck) AS ha_stck, SUM(montaza) AS montaza, "
        "MIN(datum) AS datum_od, MAX(datum) AS datum_do "
        "FROM project_daily" + where + " GROUP BY projektname", tuple(params))]
    return jsonify({"totals": rows})


@app.route("/api/projects/sync", methods=["POST"])
@api_login_required
def api_projects_sync():
    t = threading.Thread(target=sync_projects_from_azure, daemon=True)
    t.start()
    t.join(timeout=60)
    return jsonify(_sync_state)


@app.route("/api/history")
@api_login_required
def api_history():
    """Spojena historija za POP ili DP: trajni audit_log + seg_history termina."""
    entity = request.args.get("entity", "dp")
    eid = request.args.get("id", default=0, type=int)
    limit = min(request.args.get("limit", default=150, type=int), 500)
    d = db()
    events = []
    if entity == "pop":
        dp_ids = [r["id"] for r in d.execute(
            "SELECT id FROM dps WHERE pop_id=%s", (eid,))]
        for r in d.execute(
                "SELECT * FROM audit_log WHERE entity='pop' AND entity_id=%s "
                "ORDER BY id DESC LIMIT %s", (eid, limit)):
            events.append({**dict(r), "kind": "audit"})
    else:
        dp_ids = [eid]
    if dp_ids:
        ph = ",".join(["%s"] * len(dp_ids))
        for r in d.execute(
                f"SELECT * FROM audit_log WHERE entity='dp' AND entity_id IN ({ph}) "
                "ORDER BY id DESC LIMIT %s", (*dp_ids, limit)):
            events.append({**dict(r), "kind": "audit"})
        for r in d.execute(
                'SELECT h.ts, h."user", h.polje, h.vrijednost, t.aktivnost, '
                " d2.naziv AS dp_naziv "
                "FROM seg_history h JOIN segments s ON s.id=h.seg_id "
                "JOIN tasks t ON t.id=s.task_id JOIN dps d2 ON d2.id=t.dp_id "
                f"WHERE t.dp_id IN ({ph}) ORDER BY h.id DESC LIMIT %s",
                (*dp_ids, limit)):
            events.append({**dict(r), "kind": "seg"})
    events.sort(key=lambda e: e["ts"], reverse=True)
    return jsonify({"events": events[:limit]})


@app.route("/api/stats")
@api_login_required
def api_stats():
    d = db()
    by_status = {r["s"]: r["n"] for r in d.execute(
        "SELECT status s, COUNT(*) n FROM segments GROUP BY status")}
    by_odjel = [dict(r) for r in d.execute(
        "SELECT t.odjel, "
        " SUM(CASE WHEN s.status='završeno' THEN 1 ELSE 0 END) zavrseno, "
        " SUM(CASE WHEN s.status='otvoreno' THEN 1 ELSE 0 END) otvoreno "
        "FROM segments s JOIN tasks t ON t.id=s.task_id GROUP BY t.odjel ORDER BY t.odjel")]
    per_dp = [dict(r) for r in d.execute(
        "SELECT d.id, d.pop, d.naziv, d.hp, d.ha, COUNT(s.id) ukupno, "
        " SUM(CASE WHEN s.status='završeno' THEN 1 ELSE 0 END) zavrseno, "
        " SUM(CASE WHEN s.eskalacija=1 THEN 1 ELSE 0 END) eskalacije "
        "FROM dps d LEFT JOIN tasks t ON t.dp_id=d.id "
        "LEFT JOIN segments s ON s.task_id=t.id GROUP BY d.id ORDER BY d.pop, d.naziv")]
    return jsonify({"by_status": by_status, "by_odjel": by_odjel, "per_dp": per_dp})


def _to_date(s):
    """'YYYY-MM-DD' ili ISO timestamp -> datetime.date (za prave datume u Excelu)."""
    if not s:
        return None
    try:
        return datetime.date.fromisoformat(str(s)[:10])
    except (ValueError, TypeError):
        return None


# Definicije kolona: (Zaglavlje, ključ, tip)  tip: t=tekst, i=cijeli broj, f=decimalni, d=datum
TERMINI_COLS = [
    ("Projekt", "projekt", "t"), ("Kunde", "kunde", "t"), ("Projectcode", "projectcode", "t"),
    ("POP/FCP ID", "pop", "t"), ("DP", "dp", "t"), ("Lokacija", "lokacija", "t"),
    ("Voditelj", "voditelj", "t"), ("HP", "dp_hp", "i"), ("HA", "dp_ha", "i"),
    ("Aktivnost", "aktivnost", "t"), ("Odjel", "odjel", "t"), ("Status", "status", "t"),
    ("Od", "datum_od", "d"), ("Do", "datum_do", "d"), ("Trajanje (dana)", "trajanje", "i"),
    ("Kasni (dana)", "kasni", "i"), ("Eskalacija", "eskalacija_txt", "t"),
    ("Datum eskalacije", "esk_datum", "d"), ("Razlog eskalacije", "esk_razlog", "t"),
    ("Razlog kašnjenja", "kasni_razlog", "t"), ("Komentar", "komentar", "t"),
    ("Kreirao", "created_by", "t"), ("Originalno od", "orig_od", "d"),
    ("Originalno do", "orig_do", "d"), ("Pomjereno", "pomjereno", "t"),
]
DP_COLS = [
    ("Projekt", "projekt", "t"), ("Kunde", "kunde", "t"), ("POP/FCP ID", "pop", "t"),
    ("DP", "dp", "t"), ("Lokacija", "lokacija", "t"), ("Voditelj", "voditelj", "t"),
    ("HP", "hp", "i"), ("HA", "ha", "i"), ("Vlasnik", "vlasnik", "t"),
    ("Termina", "termina", "i"), ("Završeno", "zavrseno", "i"), ("% završeno", "pct", "i"),
    ("Eskalacije", "esk", "i"), ("Kasni termina", "kasni", "i"), ("Rok (Aktivacije)", "rok", "d"),
]
POP_COLS = [
    ("Projekt", "projekt", "t"), ("Kunde", "kunde", "t"), ("POP/FCP ID", "naziv", "t"),
    ("RFA", "rfa", "d"), ("Broj DP-ova", "dp_count", "i"),
    ("Kreirao", "created_by", "t"), ("Kreirano", "created_at", "d"),
]
PROJ_COLS = [
    ("Projekt", "projektname", "t"), ("Kunde", "kunde", "t"), ("Projectcode", "projectcode", "t"),
    ("Vlasnik", "vlasnik", "t"), ("HP", "hp", "f"), ("Trasa (m)", "trasa_m", "f"),
    ("HA (m)", "ha_m", "f"), ("HA kom", "ha_stck", "f"), ("Montaža", "montaza", "f"),
    ("Datum od", "datum_od", "d"), ("Datum do", "datum_do", "d"), ("Zadnji sync", "synced_at", "d"),
]
KOM_COLS = [
    ("Datum", "ts", "d"), ("Korisnik", "korisnik", "t"), ("Projekt", "projekt", "t"),
    ("POP/FCP ID", "pop", "t"), ("DP", "dp", "t"), ("Komentar", "tekst", "t"),
]


def _termini_rows():
    """Svi termini s punim kontekstom (projekt/kunde/DP) + izvedene kolone."""
    rows = [dict(r) for r in db().execute(
        "SELECT COALESCE(NULLIF(p.projekt,''), d.projekt) AS projekt, "
        "       pr.kunde AS kunde, pr.projectcode AS projectcode, "
        "       d.id AS dp_id, d.pop AS pop, d.naziv AS dp, d.lokacija AS lokacija, "
        "       d.voditelj AS voditelj, d.hp AS dp_hp, d.ha AS dp_ha, "
        "       t.aktivnost AS aktivnost, t.odjel AS odjel, s.status AS status, "
        "       s.datum_od, s.datum_do, s.orig_od, s.orig_do, s.eskalacija, "
        "       s.esk_datum, s.esk_razlog, s.kasni_razlog, s.komentar, s.created_by "
        "FROM segments s JOIN tasks t ON t.id=s.task_id JOIN dps d ON d.id=t.dp_id "
        "LEFT JOIN pops p ON p.id=d.pop_id "
        "LEFT JOIN projects pr ON pr.projektname=COALESCE(NULLIF(p.projekt,''), d.projekt) "
        "ORDER BY projekt, d.pop, d.naziv, t.id, s.datum_od")]
    today = datetime.date.today()
    for r in rows:
        od, do = _to_date(r["datum_od"]), _to_date(r["datum_do"])
        r["trajanje"] = ((do - od).days + 1) if (od and do and do >= od) else None
        r["kasni"] = (today - do).days if (do and r["status"] != "završeno" and do < today) else None
        r["eskalacija_txt"] = "da" if r["eskalacija"] else "ne"
        r["pomjereno"] = ("da" if (r.get("orig_od") and
                          (r["orig_od"] != r["datum_od"] or r["orig_do"] != r["datum_do"])) else "ne")
    return rows


def _xlsx_sheet(wb, title, table_name, columns, rows):
    """Jedan list: stilizovano zaglavlje, pravi tipovi (datumi/brojevi), Excel tabela
    (filter + naizmjenične trake), zamrznut prvi red, automatska širina kolona."""
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    ws = wb.create_sheet(title)
    ws.append([c[0] for c in columns])
    for ri, row in enumerate(rows, start=2):
        for ci, (_, key, kind) in enumerate(columns, start=1):
            v = row.get(key)
            cell = ws.cell(row=ri, column=ci)
            if kind == "d":
                dv = _to_date(v)
                if dv is not None:
                    cell.value = dv
                    cell.number_format = "DD.MM.YYYY"
            elif kind == "i":
                if v not in (None, "", "None"):
                    try: cell.value = int(float(v))
                    except (ValueError, TypeError): cell.value = v
            elif kind == "f":
                if v not in (None, ""):
                    try: cell.value = round(float(v), 2)
                    except (ValueError, TypeError): cell.value = v
                    cell.number_format = "#,##0.##"
            else:
                cell.value = "" if v is None else str(v)
    for j in range(1, len(columns) + 1):
        c = ws.cell(row=1, column=j)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
    # automatska širina iz zaglavlja + uzorka vrijednosti
    for j, (head, key, _) in enumerate(columns, start=1):
        mx = len(str(head))
        for row in rows[:300]:
            v = row.get(key)
            mx = max(mx, len(str(v)) if v not in (None, "") else 0)
        ws.column_dimensions[get_column_letter(j)].width = min(max(mx + 2, 9), 46)
    last = get_column_letter(len(columns))
    ref = "A1:%s%d" % (last, len(rows) + 1)
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    return ws


@app.route("/export/csv")
@login_required
def export_csv():
    """Glavna tabela (termini) sa svim kolonama; datumi u ISO formatu (Excel ih prepozna)."""
    rows = _termini_rows()
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow([c[0] for c in TERMINI_COLS])
    for r in rows:
        w.writerow(["" if r.get(k) is None else r.get(k) for (_, k, _) in TERMINI_COLS])
    data = io.BytesIO(("﻿" + buf.getvalue()).encode("utf-8"))
    return send_file(data, as_attachment=True, mimetype="text/csv",
                     download_name="dp_planiranje_" + datetime.date.today().isoformat() + ".csv")


@app.route("/export/xlsx")
@login_required
def export_xlsx():
    """Robustan Excel izvještaj: više listova (Termini, DP-ovi, POP-ovi, Projekti,
    Komentari) — pravi datumi/brojevi, filteri i zamrznuta zaglavlja na svakom."""
    from openpyxl import Workbook
    d = db()
    termini = _termini_rows()
    claims = {r["projektname"]: r for r in d.execute(
        "SELECT projektname, owner_name, owner_email FROM project_claims")}

    # DP-ovi (svi, uklj. bez termina) + agregati iz termina
    dps = [dict(r) for r in d.execute(
        "SELECT d.id, COALESCE(NULLIF(p.projekt,''), d.projekt) AS projekt, pr.kunde AS kunde, "
        "       d.pop AS pop, d.naziv AS dp, d.lokacija AS lokacija, d.voditelj AS voditelj, "
        "       d.hp AS hp, d.ha AS ha "
        "FROM dps d LEFT JOIN pops p ON p.id=d.pop_id "
        "LEFT JOIN projects pr ON pr.projektname=COALESCE(NULLIF(p.projekt,''), d.projekt) "
        "ORDER BY projekt, d.pop, d.naziv")]
    agg = {}
    for r in termini:
        a = agg.setdefault(r["dp_id"], {"termina": 0, "zavrseno": 0, "esk": 0, "kasni": 0, "rok": None})
        a["termina"] += 1
        if r["status"] == "završeno":
            a["zavrseno"] += 1
        if r["eskalacija"]:
            a["esk"] += 1
        if r["kasni"]:
            a["kasni"] += 1
        if "aktivacij" in (r["aktivnost"] or "").lower() and r["datum_do"]:
            if not a["rok"] or r["datum_do"] > a["rok"]:
                a["rok"] = r["datum_do"]
    for dp in dps:
        a = agg.get(dp["id"], {})
        dp["termina"] = a.get("termina", 0)
        dp["zavrseno"] = a.get("zavrseno", 0)
        dp["esk"] = a.get("esk", 0)
        dp["kasni"] = a.get("kasni", 0)
        dp["rok"] = a.get("rok")
        dp["pct"] = round(a["zavrseno"] / a["termina"] * 100) if a.get("termina") else 0
        c = claims.get(dp["projekt"])
        dp["vlasnik"] = (c["owner_name"] or c["owner_email"]) if c else ""

    # POP-ovi (uklj. one bez DP-a)
    pops = [dict(r) for r in d.execute(
        "SELECT p.id, p.projekt AS projekt, pr.kunde AS kunde, p.naziv AS naziv, "
        "       p.hp AS hp, p.ha AS ha, p.rfa AS rfa, "
        "       p.created_by AS created_by, p.created_at AS created_at, "
        "       COUNT(dd.id) AS dp_count "
        "FROM pops p LEFT JOIN dps dd ON dd.pop_id=p.id "
        "LEFT JOIN projects pr ON pr.projektname=p.projekt "
        "GROUP BY p.id, pr.kunde ORDER BY p.projekt, p.naziv")]

    # Projekti (Azure sync) + vlasnik
    projects = [dict(r) for r in d.execute(
        "SELECT projektname, kunde, projectcode, hp, trasa_m, ha_m, ha_stck, montaza, "
        "       datum_od, datum_do, synced_at FROM projects ORDER BY kunde, projektname")]
    for pr in projects:
        c = claims.get(pr["projektname"])
        pr["vlasnik"] = (c["owner_name"] or c["owner_email"]) if c else ""

    # Komentari
    comments = [dict(r) for r in d.execute(
        'SELECT c.ts AS ts, c."user" AS korisnik, c.tekst AS tekst, '
        "       d.pop AS pop, d.naziv AS dp, "
        "       COALESCE(NULLIF(p.projekt,''), d.projekt) AS projekt "
        "FROM dp_comments c JOIN dps d ON d.id=c.dp_id LEFT JOIN pops p ON p.id=d.pop_id "
        "ORDER BY c.ts DESC")]

    wb = Workbook()
    wb.remove(wb.active)   # ukloni prazni default list
    _xlsx_sheet(wb, "Termini", "Termini", TERMINI_COLS, termini)
    _xlsx_sheet(wb, "DP-ovi", "DPovi", DP_COLS, dps)
    _xlsx_sheet(wb, "POP-ovi", "POPovi", POP_COLS, pops)
    _xlsx_sheet(wb, "Projekti", "Projekti", PROJ_COLS, projects)
    _xlsx_sheet(wb, "Komentari", "Komentari", KOM_COLS, comments)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True,
                     download_name="dp_planiranje_" + datetime.date.today().isoformat() + ".xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


database.init_db(permanent_admin=PERMANENT_ADMIN)


def _sync_loop():
    """Automatski sync iz Azure: odmah pri startu pa svakih 30 minuta
    (ručno Sync dugme je uklonjeno iz UI-ja)."""
    while True:
        sync_projects_from_azure()
        time.sleep(30 * 60)


def _digest_loop():
    """Sedmični e-mail pregled: ponedjeljkom od 07h, tačno jednom
    (advisory lock + meta zapis štite od duplikata kroz više workera)."""
    while True:
        try:
            now = datetime.datetime.now()
            if now.weekday() == 0 and now.hour >= 7:
                today = now.date().isoformat()
                con = database.connect()
                cur = con.cursor()
                cur.execute("SELECT pg_try_advisory_lock(727274002)")
                if cur.fetchone()[0]:
                    cur.execute("SELECT value FROM meta WHERE key='digest_sent'")
                    r = cur.fetchone()
                    if not r or r[0] != today:
                        recipients = _emails()
                        if not recipients:
                            print("[MAIL] digest: nema primalaca (allowed_users prazna) — preskačem.")
                        # zabilježi POKUŠAJ prije SMTP-a -> nema ponovnog slanja istog dana
                        # ni kad SMTP tiho zakaže (inače loop spamuje svakih 30 min)
                        cur.execute(
                            "INSERT INTO meta(key,value) VALUES('digest_sent',%s) "
                            "ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value", (today,))
                        con.commit()
                        if recipients and not send_mail(
                                recipients, "DP Planiranje — sedmični pregled", _digest_html()):
                            print("[MAIL] digest: slanje nije uspjelo (vidi gore).")
                    cur.execute("SELECT pg_advisory_unlock(727274002)")
                con.close()
        except Exception as e:
            print("[MAIL] digest loop:", e)
        time.sleep(30 * 60)


threading.Thread(target=_sync_loop, daemon=True).start()
threading.Thread(target=_digest_loop, daemon=True).start()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(host=os.environ.get("HOST", "127.0.0.1"), port=port, debug=False,
            threaded=True)
